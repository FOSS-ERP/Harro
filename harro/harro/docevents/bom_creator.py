import frappe
from openpyxl import load_workbook

@frappe.whitelist()
def execute_uploaded_file(file_path, bom_c):
    frappe.enqueue(
            extract_bom_item_data, file_path=file_path, bom_c=bom_c, queue="long", enqueue_after_commit=True
        )
    
@frappe.whitelist()
def extract_bom_item_data(file_path, bom_c):
    doc = frappe.get_doc("BOM Creator", bom_c)

    if not file_path:
        frappe.throw("Please upload a file first.")

    # Read Excel
    if "private" in file_path:
        file_path = file_path.replace("/private", '')
        public_file_path = frappe.get_site_path("private", file_path.lstrip("/"))
    else:
        public_file_path = frappe.get_site_path("public", file_path.lstrip("/"))
    workbook = load_workbook(filename=public_file_path, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1] if cell.value]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        data.append(row_dict)

    create_item_group()

    # Map to track FG item â†’ child table row idx
    fg_row_map = {}

    for row in data:
        artikel = row.get("Artikel")
        baugruppe = row.get("Baugruppe")
        menge = row.get("Menge")
        mengeneinheit = row.get("Mengeneinheit")
        StrukturklasseKopf = row.get("StrukturklasseKopf")
        StrukturklassePos = row.get("StrukturklassePos")

        uom = None
        if mengeneinheit:
            if stock_uom := frappe.db.exists("UOM", {"custom_german_uom" : mengeneinheit}):
                uom = stock_uom

        # Create Item if it doesn't exist
        if not frappe.db.exists("Item", artikel, ):
            create_item(artikel, row, uom, item_group=StrukturklassePos)
        if not frappe.db.exists("Item", baugruppe ):
            create_item(baugruppe, row, uom=None, item_group=StrukturklasseKopf)

        new_row = {
            "item_code": artikel,
            "fg_item": baugruppe,
            "qty" : menge
        }

        # Set parent_row_no only if FG item already exists in table
        if baugruppe in fg_row_map:
            new_row["parent_row_no"] = fg_row_map[baugruppe]
        else:
            # Top-level FG item (first occurrence)
            new_row["fg_reference_id"] = doc.name
            new_row["fg_item"] = doc.item_code

        # Append row and store child table index
        doc.append("items", new_row)
        fg_row_map[artikel] = doc.items[-1].idx  # map this item's row index

    doc.save()


def create_item(item, row, uom=None, item_group="Production Part"):

    if item_group_ := frappe.db.exists("Item Group", {"custom_german_name_of_item_group" : item_group}):
       item_group = item_group_
    else:
        frappe.get_doc(
            {
                "item_group_name": item_group,
                "custom_german_name_of_item_group" : item_group,
                "parent_item_group": "All Item Groups",
                "old_parent": "All Item Groups",
                "doctype": "Item Group"
            }
        ).insert()

    item_doc = frappe.get_doc({
                "doctype" : "Item",
                "item_code" : item,
                "item_group" : item_group,
                "stock_uom" : uom or "Nos",
                "valuation_rate" : 0
            })

    # if row.get("cHerstellername"):
    #     if frappe.db.exists("Item Manufacturer", { "item_code" : item, "manufacturer" :  row.get("cHerstellername")}):
    #         item_doc.custom_manufacturer = row.get("cHerstellername")
    
    labels = [
        "Baugruppe",
        "Baugruppe HH_India",
        "StrukturklasseKopf",
        "Revision",
        "Revision alt",
        "Position",
        "Position alt",
        "Artikel",
        "Artikel HH_India",
        "StrukturklassePos",
        "Revision Artikel",
        "Artikel alt",
        "Artikel alt HH_India",
        "StrukturklassePos alt",
        "Revision Artikel alt",
        "Menge",
        "Mengeneinheit",
        "Menge alt",
        "Mengeneinheit alt",
        "initial",
        "status",
        "cHerstellernr",
        "cHerstellerBez",
        "cHerstellerBez2",
        "cHerstellerBez3",
        "cHerstellerBez4",
        "cBestellnummer",
        "iLieferant",
        "cLieferantSuchbegriff",
        "cLiefBez",
        "cLiefBez2",
        "cLiefBez3",
        "cLiefBez4",
        "cLiefBestellnummer",
        "Baugruppe Bez1",
        "Baugruppe Bez2",
        "Baugruppe Bez3",
        "Baugruppe Bez4",
        "Baugruppe pruefplan",
        "Baugruppe Werkstoff",
        "Baugruppe Oberflaechenbehandlung",
        "Baugruppe Rohteil",
        "Baugruppe Schweissteil",
        "Baugruppe Halbzeug",
        "Artikel Bez1",
        "Artikel Bez2",
        "Artikel Bez3",
        "Artikel Bez4",
        "Artikel pruefplan",
        "Artikel Werkstoff",
        "Artikel Oberflaechenbehandlung",
        "Artikel Rohteil",
        "Artikel Schweissteil",
        "Artikel Halbzeug",
        "Stufe Baugruppenpos",
        "HerstellerNr HH_India",
        "sequenz",
        "Anzahl",
        "Menge Brutto",
        "Gesamtmenge"
    ]

    for l in labels:
        fieldname = make_fieldname(l)
        item_doc.update({
            fieldname : row.get(l)
        })
    item_doc.insert(ignore_permissions=True)

    
def create_item_group():
    if not frappe.db.exists("Item Group", "Production Part"):
        frappe.get_doc({
            "doctype" : "Item Group",
            "parent_item_group" : "All Item Groups",
            "item_group_name" : "Production Part"
        }).insert()

def make_fieldname(label):
    return label.strip().lower().replace(" ", "_") 
    